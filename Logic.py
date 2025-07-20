import requests
import json
import yaml
from openpyxl import Workbook, styles

# === Swagger to Excel Core Logic ===
def fetch_from_swagger(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
    except requests.RequestException as e:
        return None, f"❌ Error fetching URL:\n{e}"

    raw_data = response.text
    try:
        if 'json' in response.headers.get('Content-Type', '') or url.endswith('.json'):
            data = json.loads(raw_data)
        elif url.endswith(('.yaml', '.yml')) or 'yaml' in response.headers.get('Content-Type', ''):
            data = yaml.safe_load(raw_data)
        else:
            try:
                data = json.loads(raw_data)
            except json.JSONDecodeError:
                data = yaml.safe_load(raw_data)
    except Exception as e:
        return None, f"❌ Error parsing file:\n{e}"

    return data

def extract_swagger_data(swagger):
    endpoint = swagger.get('paths', {})
    schema = swagger.get('components', {}).get('schemas', {})
    return endpoint, schema

def resolve_ref(ref, root):
    keys = ref.lstrip("#/").split("/")
    for key in keys:
        root = root[key]
    return root

def generate_payload(schema):
    if "type" not in schema:
        return {}

    if schema["type"] == "object":
        return {prop: generate_payload(schema["properties"][prop]) for prop in schema.get("properties", {})}
    elif schema["type"] == "array":
        return [generate_payload(schema["items"])]
    elif schema["type"] == "string":
        return schema.get("example", "string")
    elif schema["type"] == "integer":
        return schema.get("example", 0)
    elif schema["type"] == "boolean":
        return schema.get("example", False)
    return None

def convert_to_excel(swagger_data):
    endpoints, schemas = extract_swagger_data(swagger_data)
    workbook = Workbook()

    # Sheet_1
    sheet1 = workbook.active
    sheet1.title = "API Endpoints"
    srNo=1
    sheet1.append(["Sr. No.", "Module", "Request Type", "Endpoints", "Payloads", "Summary", "Parameters"])

    sheet1.column_dimensions['A'].width = 8
    sheet1.column_dimensions['B'].width = 20
    sheet1.column_dimensions['C'].width = 15
    sheet1.column_dimensions['D'].width = 35
    sheet1.column_dimensions['E'].width = 30
    sheet1.column_dimensions['F'].width = 30
    sheet1.column_dimensions['G'].width = 30

    sheet1.row_dimensions[1].height = 30

    header_fill = styles.PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = styles.Font(bold=True)
    center_alignment = styles.Alignment(horizontal="center", vertical="center")
    thin_border = styles.Border(
        left=styles.Side(style='thin', color='000000'),
        right=styles.Side(style='thin', color='000000'),
        top=styles.Side(style='thin', color='000000'),
        bottom=styles.Side(style='thin', color='000000')
    )

    for cell in sheet1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    for endpoints, requestTypes in endpoints.items():
        for requestType, details in requestTypes.items():
            tags = details.get("tags", ["Unknown"])
            module = tags[0] if tags else "Unknown"

            requestBody = details.get("requestBody", {}).get("content", {}).get("application/json", {})
            ref = requestBody.get("schema", {}).get("$ref")
            payload = ""
            try:
                if ref:
                    schema = resolve_ref(ref, swagger_data)
                    generated_payload  = generate_payload(schema)
                    payload = json.dumps(generated_payload, indent=2)
                else:
                    payload = "No Body"
            except Exception as e:
                payload = f"❌ Unable to Fetch Payload + {e}"

            summary = details.get('summary', '')
            parameters = details.get('parameters', [])
            param_str = ', '.join([f"{p.get('name')}({p.get('in')})" for p in parameters])

            sheet1.append([srNo,module.capitalize(),requestType.capitalize(), endpoints, payload, summary, param_str])
            srNo += 1
    
    for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
        for cell in row:
            cell.border = thin_border

    for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row):
        for col_index in [1, 2, 3]: 
            row[col_index - 1].alignment = center_alignment

    # Sheet_2
    # sheet2 = workbook.create_sheet("Schemas")
    # sheet2.append(["Schema Name", "Properties"])

    # for name, schema in schemas.items():
    #     props = schema.get('properties', {})
    #     prop_details = ', '.join([f"{p}: {v.get('type', 'object')}" for p, v in props.items()])
    #     sheet2.append([name, prop_details])

    return workbook

